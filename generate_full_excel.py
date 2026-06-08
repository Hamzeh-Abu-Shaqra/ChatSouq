"""
Generates: ChatSouq_Jordan_Data_FULL.xlsx
Every row of every data file, 20+ sheets.
"""

import json, csv, io, os, sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE   = Path("/Users/Hamzeh23/Desktop/chatsouq/data")
OUT    = Path.home() / "Desktop" / "ChatSouq_Jordan_Data_FULL.xlsx"
RAW    = BASE / "knowledge-graph" / "raw"
KG_OUT = BASE / "knowledge-graph" / "output"
SEED   = BASE / "seed"

# ── Style helpers ─────────────────────────────────────────────────────────────

HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALT_FILL     = PatternFill("solid", fgColor="EBF3FB")
BORDER_SIDE  = Side(style="thin", color="BDD7EE")
CELL_BORDER  = Border(bottom=BORDER_SIDE)

TIER_COLORS = {
    "luxury":    "F4D03F",
    "Luxury":    "F4D03F",
    "upscale":   "82E0AA",
    "Upscale":   "82E0AA",
    "mid-range": "85C1E9",
    "Mid-Range": "85C1E9",
    "budget":    "F1948A",
    "Budget":    "F1948A",
}

def style_header(ws, row=1):
    for cell in ws[row]:
        cell.font  = HEADER_FONT
        cell.fill  = HEADER_FILL
        cell.alignment = HEADER_ALIGN

def style_rows(ws, start=2, tier_col=None):
    for i, row in enumerate(ws.iter_rows(min_row=start)):
        fill = ALT_FILL if i % 2 == 0 else None
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if fill:
                cell.fill = fill
            cell.border = CELL_BORDER
        if tier_col:
            tier_cell = row[tier_col - 1]
            color = TIER_COLORS.get(str(tier_cell.value), None)
            if color:
                tier_cell.fill = PatternFill("solid", fgColor=color)
                tier_cell.font = Font(bold=True)

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def freeze(ws, cell="A2"):
    ws.freeze_panes = cell

def write_sheet(ws, headers, rows, widths=None, tier_col=None):
    ws.append(headers)
    style_header(ws)
    for row in rows:
        ws.append(row)
    style_rows(ws, tier_col=tier_col)
    freeze(ws)
    if widths:
        set_col_widths(ws, widths)
    # Auto-filter
    ws.auto_filter.ref = ws.dimensions

def write_sheet_from_dicts(ws, dicts, widths=None, tier_col=None, key_order=None):
    if not dicts:
        ws.append(["(no data)"])
        return
    headers = key_order if key_order else list(dicts[0].keys())
    rows = [[str(d.get(h, "") or "") for h in headers] for d in dicts]
    write_sheet(ws, headers, rows, widths, tier_col)

# ── Load helpers ──────────────────────────────────────────────────────────────

def load_ndjson(path):
    rows = []
    with open(path) as f:
        for line in f:
            line = line.strip()
            if line:
                rows.append(json.loads(line))
    return rows

def load_json_list(path, *keys):
    with open(path) as f:
        d = json.load(f)
    if isinstance(d, list):
        return d
    for k in keys:
        if k in d:
            return d[k]
    # Try common key names
    for k in ("businesses","places","institutions","items","entities","data"):
        if k in d:
            return d[k]
    return []

def load_csv(path):
    rows = []
    with open(path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            rows.append(dict(row))
    return rows

def flatten(val):
    if isinstance(val, list):
        return "; ".join(str(v) for v in val)
    if isinstance(val, dict):
        return json.dumps(val, ensure_ascii=False)
    return str(val) if val is not None else ""

def normalise_row(d):
    return {k: flatten(v) for k, v in d.items()}

# ── Workbook ──────────────────────────────────────────────────────────────────

wb = Workbook()
wb.remove(wb.active)  # remove default sheet

# ════════════════════════════════════════════════════════════════════
# SHEET 1 – OVERVIEW
# ════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Overview")
from datetime import date
overview = [
    ["ChatSouq – Complete Jordan Data Export", ""],
    ["Generated", str(date.today())],
    ["", ""],
    ["METRIC", "VALUE"],
    ["Total Product Listings (listings.ndjson)", "62,898"],
    ["Products Supplement", "14"],
    ["Vendors", "64"],
    ["Knowledge Graph Entities (Master.csv)", "8,306"],
    ["OSM Places (places.json)", "8,350"],
    ["Automotive, Retail & NGO (raw)", "203"],
    ["Business Directories (raw)", "270"],
    ["E-Commerce Products (raw)", "502"],
    ["Entertainment & Sports (raw)", "203"],
    ["Financial Services (raw)", "130"],
    ["Government Institutions (raw)", "139"],
    ["Restaurants & E-Commerce (raw)", "269"],
    ["Technology Sector (raw)", "171"],
    ["Tourism, Healthcare & Govt (raw)", "163"],
    ["Amman Neighborhoods Documented", "17"],
    ["Jordan Governorates", "12"],
    ["Jordan Cities", "8"],
    ["Product Categories (Canonical)", "23"],
    ["Place Category Types (OSM-normalised)", "78"],
    ["Government Institutions (curated)", "48"],
    ["Tourism Attractions (curated)", "15"],
    ["Major Corporations (curated)", "33"],
    ["", ""],
    ["SHEET", "DESCRIPTION"],
    ["Overview", "This sheet – stats & sheet guide"],
    ["Amman Neighborhoods", "17 districts: tier, rent range JOD, characteristics, pros, cons, best-for"],
    ["Jordan Governorates", "12 governorates with lat/lng centroids and notes"],
    ["Jordan Cities", "8 major cities with Arabic names and role"],
    ["Product Categories", "23 canonical departments + Arabic equivalents"],
    ["Vendors", "64 verified Jordan vendors with website & social"],
    ["Product Listings", "All 62,898 product listings (full data)"],
    ["Products Supplement", "14 additional curated products"],
    ["Listings – Category Summary", "62,898 listings broken down by raw category"],
    ["Knowledge Graph – Master", "All 8,306 entities: services, food, tourism, healthcare, education…"],
    ["OSM Places (Raw)", "8,350 OpenStreetMap places across Jordan"],
    ["Automotive, Retail & NGO", "203 entries: car dealers, NGOs, retail chains"],
    ["Business Directories", "270 curated business entries"],
    ["E-Commerce Products", "502 curated e-commerce product entries"],
    ["Entertainment & Sports", "203 venues, clubs, sport facilities"],
    ["Financial Services", "130 banks, insurance, money exchange"],
    ["Government Institutions", "139 ministries, departments, public authorities"],
    ["Restaurants & E-Commerce", "269 restaurants and online retail"],
    ["Technology Sector", "171 tech companies, startups, accelerators"],
    ["Tourism, Healthcare & Govt", "163 tourism sites, hospitals, government offices"],
    ["Place Categories (OSM)", "78 normalised place types by OSM key"],
    ["Govt Institutions – Curated", "48 key institutions with sector and Arabic name"],
    ["Tourism Attractions – Curated", "15 major sites with UNESCO status, fees, tips"],
    ["Major Corporations – Curated", "33 key Jordanian companies across all sectors"],
]
for row in overview:
    ws.append(row)
ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
ws["A4"].font = Font(bold=True)
ws["A27"].font = Font(bold=True)
set_col_widths(ws, [42, 64])

print("Sheet 1/24 – Overview ✓")

# ════════════════════════════════════════════════════════════════════
# SHEET 2 – AMMAN NEIGHBORHOODS
# ════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Amman Neighborhoods")
neighborhoods = [
    # name, name_ar, tier, rent_min, rent_max, area, characteristics, pros, cons, best_for, price_level
    ("Dabouq","دابوق","Luxury",1000,3500,"Far West Amman",
     "Quiet villa district; Far west Amman; Lush greenery; Very private",
     "Most spacious & private; Large villas with gardens; Very low density; Excellent air quality",
     "Far from city center (30+ min); Car absolutely essential; Limited nearby shops",
     "High-budget families; Executives; Those wanting maximum space & privacy","$$$$"),
    ("Abdoun","عبدون","Luxury",900,2500,"West Amman – 5th Circle",
     "Embassy district; Tree-lined streets; Villas & upscale apartments; Near 5th Circle",
     "Prestige address; Extremely safe & secure; Near best restaurants in Amman; International community",
     "Car essential; Very high cost of living; Limited public transit",
     "Expats; Senior executives; Diplomatic families","$$$$"),
    ("Deir Ghbar","دير غبار","Luxury",800,2000,"West Amman – 5th Circle",
     "Quiet hilltop; Upscale residential; Private community feel; Near 5th Circle",
     "Panoramic city views; Very secure; Modern buildings; Close to Abdoun amenities",
     "Car essential; Limited walkability; Higher service charges",
     "Families; Expats; Professionals seeking quiet luxury","$$$$"),
    ("Rabieh","الرابية","Upscale",700,1800,"West Amman – 4th/5th Circles",
     "Quiet suburban; Upscale housing; Near 4th & 5th circles; Green spaces",
     "Very peaceful & quiet; High-quality buildings; Good international schools nearby; Safe & secure",
     "Need a car; Limited walkability; Fewer cafes & shops",
     "Families with children; Expats; Professionals","$$$"),
    ("Um Uthaina","أم أذينة","Upscale",550,1200,"West Amman – 4th Circle",
     "Residential & commercial mix; Near 4th Circle; Well-connected; Many amenities",
     "Central-west location; Walkable to shops; Good schools & hospitals; Less traffic than Sweifieh",
     "Getting denser; Moderate traffic; Mixed commercial-residential",
     "Young professionals; Couples; Families who want convenience","$$$"),
    ("Sweifieh","الصويفية","Upscale",500,1100,"West Amman – Central",
     "Major shopping district; Cafes & restaurants; Active nightlife; Central-west Amman",
     "Everything within walking distance; Excellent transport; Great dining & entertainment; Lively atmosphere",
     "Heavy traffic congestion; Noisy; Parking nightmare; Higher rents for size",
     "Young singles & couples; Expats who like city energy; Those who don't need a car","$$$"),
    ("Shmeisani","الشميساني","Mid-Range",450,900,"Central Amman",
     "Business & financial district; Banks & corporate offices; Central location; Good infrastructure",
     "Most central location in Amman; Walking distance to offices & banks; Great connectivity; Diverse amenities",
     "Commercial feel; Heavy peak-hour traffic; Less residential charm; Noisy on weekdays",
     "Business professionals; Singles; Those who work in the area","$$"),
    ("Jabal Amman","جبل عمان","Mid-Range",400,850,"Central Amman – 1st–4th Circles",
     "Historic trendy district; 1st–4th Circles; Rainbow Street; Cafes & boutiques",
     "Cultural & artsy vibe; Walkable neighbourhood; Best café & dining scene in Amman; Historic character",
     "Hilly terrain (hard on older people); Older buildings; Very limited parking",
     "Creatives & artists; Young professionals; Culture & food lovers","$$"),
    ("Wadi Saqra / 3rd Circle","وادي صقرة / الدوار الثالث","Mid-Range",350,800,"Central-West Amman",
     "Central-west Amman; Near 3rd Circle; Mix of residential & offices; Convenient location",
     "Very central; Good connectivity to all areas; Walkable to many services; Character neighbourhood",
     "Traffic at circles; Parking limited; Older building stock",
     "Young professionals; Singles; Couples who value central location","$$"),
    ("Khalda","خلدا","Mid-Range",380,780,"West Amman – University Area",
     "Established residential; Near University of Jordan; Family-oriented; Quieter pace",
     "Peaceful atmosphere; Family-friendly community; Reasonably priced for quality; Good schools & uni nearby",
     "Needs a car for most errands; Less entertainment & nightlife",
     "Families; Students at UJ; Those seeking quiet family life","$$"),
    ("Tlaa Al-Ali","تلاع العلي","Mid-Range",320,700,"West Amman – Suburbs",
     "Residential suburb; Newer developments; Family-focused; Quieter",
     "Modern apartments with good finishes; Spacious for the price; Family community feel; Less congested",
     "Further from city center (20+ min); Car essential; Less entertainment",
     "Young families; Those seeking space at reasonable cost","$$"),
    ("Sweileh","سويلح","Mid-Range",250,550,"West Amman – University Area",
     "Near University of Jordan; Active area; Mixed residential-commercial; Affordable",
     "Very affordable for west Amman; Busy local scene; Good transport options; Near university",
     "Can be chaotic; Heavy traffic on main roads; Less polished environment",
     "Students; Budget-conscious professionals; University staff","$"),
    ("Jubeiha","الجبيهة","Budget",220,480,"West Amman – University Hub",
     "University student hub; Affordable; Lively local cafes; Near UJ & JUST",
     "Lowest-cost option in west Amman; Lively student atmosphere; Many cheap eats & cafes; Near University of Jordan",
     "Noisier & busier; Less polished; Can feel overcrowded in student season",
     "University students; Young budget renters; Those on tight budgets near campus","$"),
    ("Jabal Hussein","جبل الحسين","Budget",180,400,"Central East Amman",
     "Central east Amman; Dense residential; Affordable; Local markets",
     "Very central location; Affordable rents; Good public transport; Strong local community",
     "Dense & crowded; Older buildings; Limited amenities",
     "Budget renters; Those needing central location on tight budget; Local families","$"),
    ("Jabal Al-Nuzha","جبل النزهة","Budget",150,350,"East Amman",
     "East Amman residential; Affordable; Family neighbourhood; Local community",
     "Affordable family apartments; Good community feel; Accessible public transport",
     "Limited modern amenities; Older building stock; Distance from west Amman",
     "Families on tight budgets; Local Jordanians; Those working in east Amman","$"),
    ("Marka","ماركا","Budget",130,320,"East Amman",
     "East Amman; Working-class area; Very affordable; Near airport road",
     "Cheapest rents in Amman; Near Queen Alia Highway; Strong local community",
     "Far from west Amman amenities; Older housing stock; Less infrastructure",
     "Very tight budgets; Those working near the airport or east Amman","$"),
    ("Sahab","سحاب","Budget",110,280,"South-East Amman",
     "South-east Amman suburb; Industrial area; Very affordable; Spacious",
     "Lowest rents in greater Amman area; Larger apartments for price; Quiet residential streets",
     "Very far from city center (30+ min drive); Limited local services; Far from top schools",
     "Extremely tight budgets; Those working in industrial areas; Large families needing space","$"),
]
headers = ["Neighborhood","Arabic Name","Tier","Rent Min (JOD/mo)","Rent Max (JOD/mo)",
           "Area / Location","Characteristics","Pros","Cons","Best For","Price Level"]
write_sheet(ws, headers, [list(r) for r in neighborhoods],
            widths=[22,20,11,18,18,26,52,60,48,52,12], tier_col=3)
print("Sheet 2/24 – Amman Neighborhoods ✓")

# ════════════════════════════════════════════════════════════════════
# SHEET 3 – JORDAN GOVERNORATES
# ════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Jordan Governorates")
govs = [
    (1,"Amman","عمّان","Amman",31.95,35.93,"Largest; political & economic capital; 4.2M+ pop"),
    (2,"Irbid","إربد","Irbid",32.55,35.85,"2nd most populous; northern Jordan; major university city"),
    (3,"Zarqa","الزرقاء","Zarqa",32.07,36.09,"Industrial centre; large Syrian refugee population"),
    (4,"Mafraq","المفرق","Mafraq",32.34,36.21,"Border with Syria; large Bedouin community"),
    (5,"Balqa","البلقاء","Salt (Al-Salt)",32.04,35.73,"Includes Dead Sea (Sweimeh); Salt UNESCO city"),
    (6,"Madaba","مادبا","Madaba",31.72,35.79,"Famous mosaic map; Wadi Mujib; Christian heritage"),
    (7,"Karak","الكرك","Karak",31.18,35.70,"Crusader castle; King's Highway; agricultural"),
    (8,"Tafilah","الطفيلة","Tafilah",30.84,35.60,"Dana Biosphere Reserve; relatively rural"),
    (9,"Ma'an","معان","Ma'an",30.19,35.73,"Largest by area; gateway to Petra & Wadi Rum"),
    (10,"Aqaba","العقبة","Aqaba",29.53,35.01,"Only seaport; Red Sea; ASEZA free economic zone"),
    (11,"Jerash","جرش","Jerash",32.27,35.90,"Best-preserved Roman city in Jordan; 'Pompeii of the East'"),
    (12,"Ajloun","عجلون","Ajloun",32.33,35.75,"Ajloun Castle (Saladin era); forest reserve; eco-tourism"),
]
write_sheet(ws, ["#","Governorate","Arabic Name","Capital City","Lat (°N)","Lng (°E)","Notes"],
            [list(r) for r in govs], widths=[4,14,16,18,10,10,55])
print("Sheet 3/24 – Jordan Governorates ✓")

# ════════════════════════════════════════════════════════════════════
# SHEET 4 – JORDAN CITIES
# ════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Jordan Cities")
cities = [
    ("Amman","عمّان","Amman","4,200,000+","Capital; political, economic & cultural hub of Jordan"),
    ("Irbid","إربد","Irbid","660,000+","Northern university city; Jordan's 2nd largest"),
    ("Zarqa","الزرقاء","Zarqa","600,000+","Industrial & manufacturing hub; 3rd largest city"),
    ("Aqaba","العقبة","Aqaba","148,000+","Red Sea port; tourism & ASEZA free zone"),
    ("Salt","السلط","Balqa","90,000+","UNESCO World Heritage (2021); historic Ottoman trade city"),
    ("Madaba","مادبا","Madaba","60,000+","City of mosaics; Christian heritage; tourism"),
    ("Jerash","جرش","Jerash","50,000+","Roman ruins; one of best-preserved in world"),
    ("Petra","البتراء","Ma'an","25,000+","UNESCO World Heritage; New Seven Wonder of the World"),
    ("Wadi Rum Village","رم","Ma'an","5,000+","Gateway to Wadi Rum desert; Bedouin community"),
]
write_sheet(ws, ["City","Arabic Name","Governorate","Population (est.)","Role"],
            [list(r) for r in cities], widths=[14,16,14,18,58])
print("Sheet 4/24 – Jordan Cities ✓")

# ════════════════════════════════════════════════════════════════════
# SHEET 5 – PRODUCT CATEGORIES
# ════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Product Categories")
cats = [
    (1,"Beauty & Skincare","العناية بالبشرة","Serums, moisturisers, cleansers, sunscreen, toners","beauty box jordan, gifts center, personal care"),
    (2,"Makeup","مكياج","Lipstick, foundation, eyeshadow, mascara, concealer, blush","beauty box jordan"),
    (3,"Hair Care","العناية بالشعر","Shampoo, conditioner, hair dryers, hair oil, hair serum","beauty box jordan"),
    (4,"Perfume & Fragrance","عطور","EDP, EDT, cologne, oud, body mist, gift sets","gifts center, beauty box jordan"),
    (5,"Health & Wellness","الصحة والعافية","Vitamins, supplements, dental care, oral-b, thermometers","various"),
    (6,"Mobile Phones","هواتف ذكية","iPhone, Samsung Galaxy, Xiaomi, Huawei, phone cases","multi-vendor"),
    (7,"Tablets","أجهزة لوحية","iPad, Samsung Galaxy Tab, Lenovo Tab, Huawei MatePad","multi-vendor"),
    (8,"Computers & Laptops","أجهزة كمبيوتر","MacBook, Dell, HP, Lenovo, Asus; monitors, keyboards, mice","multi-vendor"),
    (9,"Audio & Headphones","سماعات وصوتيات","Headphones, earbuds, AirPods, speakers, soundbars, headsets","multi-vendor"),
    (10,"TVs & Displays","تلفزيونات وشاشات","Smart TV, OLED, 4K, projectors, curved displays","multi-vendor"),
    (11,"Cameras","كاميرات","DSLR, mirrorless, GoPro, action cams, lenses, tripods","multi-vendor"),
    (12,"Gaming","ألعاب إلكترونية","PS5, Xbox Series, Nintendo Switch, controllers, gaming chairs","multi-vendor"),
    (13,"Electronics & Accessories","إلكترونيات وملحقات","Chargers, cables, power banks, HDMI, USB hubs, routers, SD cards","multi-vendor"),
    (14,"Home Appliances","أجهزة منزلية","Blenders, microwaves, fridges, vacuum cleaners, kettles, irons, fans","multi-vendor"),
    (15,"Home & Living","المنزل والديكور","Vases, plants, cushions, candles, lamps, curtains, bedding, rugs","vyne flower boutique"),
    (16,"Watches & Accessories","ساعات وإكسسوارات","Smartwatch, luxury watch, fashion watch, straps, wristbands","gifts center, multi-vendor"),
    (17,"Jewelry","مجوهرات","Necklaces, rings, bracelets, earrings, pendants, gold, silver","gifts center"),
    (18,"Bags & Luggage","حقائب وأمتعة","Backpacks, handbags, suitcases, trolleys, wallets, purses","multi-vendor"),
    (19,"Toys & Games","ألعاب أطفال","LEGO, dolls, remote control cars, puzzles, board games","multi-vendor"),
    (20,"Baby & Kids","مستلزمات الأطفال","Strollers, baby food, diapers, nursery decor, car seats","multi-vendor"),
    (21,"Food & Gourmet","أطعمة فاخرة","Belgian chocolates, specialty coffee, tea, honey, dates, gift boxes","neuhaus, multi-vendor"),
    (22,"Sports & Outdoors","رياضة وأماكن مفتوحة","Dumbbells, yoga mats, treadmills, tents, sleeping bags, cycles","multi-vendor"),
    (23,"Stationery & Office","قرطاسية ومكتبيات","Notebooks, pens, planners, agendas, art supplies, office supplies","multi-vendor"),
]
write_sheet(ws, ["#","Canonical Category","Arabic Equivalent","Subcategory Examples","Key Vendors / Channels"],
            [list(r) for r in cats], widths=[4,24,22,48,32])
print("Sheet 5/24 – Product Categories ✓")

# ════════════════════════════════════════════════════════════════════
# SHEET 6 – VENDORS
# ════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Vendors")
vendors = load_ndjson(SEED / "vendors.ndjson")
vendor_rows = []
for v in vendors:
    vendor_rows.append([
        v.get("id",""), v.get("businessName",""), v.get("category",""),
        v.get("description",""), v.get("location",""),
        v.get("websiteUrl",""), v.get("instagramUrl",""), v.get("status",""),
    ])
write_sheet(ws, ["ID","Business Name","Category","Description","Location","Website","Instagram","Status"],
            vendor_rows, widths=[6,30,22,65,16,36,34,10])
print(f"Sheet 6/24 – Vendors ({len(vendor_rows)} rows) ✓")

# ════════════════════════════════════════════════════════════════════
# SHEET 7 – PRODUCT LISTINGS (full 62,898)
# ════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Product Listings")
print("Loading listings (62,898 rows)…", end=" ", flush=True)
listings = load_ndjson(SEED / "listings.ndjson")
listing_rows = []
for l in listings:
    listing_rows.append([
        l.get("id",""), l.get("vendorId",""), l.get("name",""),
        l.get("category",""), l.get("brand",""),
        l.get("price",""), l.get("currency","JOD"),
        l.get("imageUrl",""), l.get("sourceUrl",""),
        (l.get("description","") or "")[:300],
    ])
write_sheet(ws, ["ID","Vendor ID","Name","Category","Brand","Price","Currency","Image URL","Source URL","Description (truncated)"],
            listing_rows, widths=[8,10,42,24,18,8,10,42,42,60])
print(f"{len(listing_rows)} rows ✓")

# ════════════════════════════════════════════════════════════════════
# SHEET 8 – PRODUCTS SUPPLEMENT
# ════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Products Supplement")
supp = load_ndjson(SEED / "products_supplement.ndjson")
if supp:
    norm = [normalise_row(r) for r in supp]
    write_sheet_from_dicts(ws, norm)
print(f"Sheet 8/24 – Products Supplement ({len(supp)} rows) ✓")

# ════════════════════════════════════════════════════════════════════
# SHEET 9 – LISTINGS CATEGORY SUMMARY
# ════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Listings – Category Summary")
from collections import Counter
counts = Counter(l.get("category","Unknown") for l in listings)
total = sum(counts.values())
summary_rows = sorted(counts.items(), key=lambda x: -x[1])
write_sheet(ws, ["Raw Category","Count","% of Total"],
            [[cat, cnt, f"{cnt/total*100:.1f}%"] for cat,cnt in summary_rows] +
            [["TOTAL", total, "100.0%"]],
            widths=[30,10,12])
print(f"Sheet 9/24 – Listings Category Summary ({len(counts)} categories) ✓")

# ════════════════════════════════════════════════════════════════════
# SHEET 10 – KNOWLEDGE GRAPH MASTER (full 8,306)
# ════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Knowledge Graph – Master")
print("Loading Knowledge Graph Master (8,306 rows)…", end=" ", flush=True)
master = load_csv(KG_OUT / "Master.csv")
write_sheet_from_dicts(ws, master,
    widths=[38,30,30,16,20,55,32,26,26,14,20,28,12,12,10,10,18,8,30,30,22,7,18,10,22,6,6,6,6,6,6])
print(f"{len(master)} rows ✓")

# ════════════════════════════════════════════════════════════════════
# SHEET 11 – OSM PLACES (full 8,350)
# ════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("OSM Places (Raw)")
print("Loading OSM Places (8,350 rows)…", end=" ", flush=True)
places = load_json_list(RAW / "places.json")
norm_places = [normalise_row(p) for p in places]
write_sheet_from_dicts(ws, norm_places,
    widths=[32,28,16,20,14,14,28,16,30,20,10,10,38,16,10])
print(f"{len(places)} rows ✓")

# ════════════════════════════════════════════════════════════════════
# SHEETS 12–19 – RAW KNOWLEDGE GRAPH SOURCES
# ════════════════════════════════════════════════════════════════════
raw_sources = [
    ("Automotive, Retail & NGO",  "automotive_retail_ngo.json",     [30,20,18,20,55,30,16,14,14,30]),
    ("Business Directories",       "business_directories.json",       [30,28,16,20,55,30,26,16,14,14,28,10,36,10]),
    ("E-Commerce Products",        "ecommerce_products.json",         [28,28,28,18,20,55,10,36,14,14,24,18]),
    ("Entertainment & Sports",     "entertainment_sports.json",       [30,22,18,20,55,30,16,14,14,30]),
    ("Financial Services",         "financial_services.json",         [30,22,18,20,55,30,16,14,14,30,18]),
    ("Government Institutions",    "jordan_government_institutions.json",[30,28,28,18,20,55,30,14,14,30,18,10]),
    ("Restaurants & E-Commerce",   "restaurants_ecommerce.json",      [30,28,16,20,55,30,26,16,14,14,10,36,10]),
    ("Technology Sector",          "technology_sector.json",          [30,22,18,20,55,30,14,14,30,10]),
    ("Tourism, Healthcare & Govt", "tourism_healthcare_govt.json",    [30,28,16,20,55,30,16,14,14,28,36,10,8,30]),
]
sheet_num = 12
for sheet_name, filename, widths in raw_sources:
    ws = wb.create_sheet(sheet_name)
    data = load_json_list(RAW / filename)
    norm = [normalise_row(r) for r in data]
    write_sheet_from_dicts(ws, norm, widths=widths)
    print(f"Sheet {sheet_num}/24 – {sheet_name} ({len(data)} rows) ✓")
    sheet_num += 1

# ════════════════════════════════════════════════════════════════════
# SHEET 21 – PLACE CATEGORIES (OSM normalised)
# ════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Place Categories (OSM)")
place_cat_data = [
    ("Amenity","amenity=restaurant","Restaurant","Food & Beverage"),
    ("Amenity","amenity=cafe","Cafe","Food & Beverage"),
    ("Amenity","amenity=fast_food","Fast Food","Food & Beverage"),
    ("Amenity","amenity=bar","Bar","Food & Beverage"),
    ("Amenity","amenity=pub","Pub","Food & Beverage"),
    ("Amenity","amenity=ice_cream","Ice Cream","Food & Beverage"),
    ("Amenity","amenity=food_court","Food Court","Food & Beverage"),
    ("Amenity","amenity=pharmacy","Pharmacy","Healthcare"),
    ("Amenity","amenity=hospital","Hospital","Healthcare"),
    ("Amenity","amenity=clinic","Clinic","Healthcare"),
    ("Amenity","amenity=doctors","Doctors","Healthcare"),
    ("Amenity","amenity=dentist","Dentist","Healthcare"),
    ("Amenity","amenity=veterinary","Veterinary","Healthcare"),
    ("Amenity","amenity=bank","Bank","Financial"),
    ("Amenity","amenity=bureau_de_change","Currency Exchange","Financial"),
    ("Amenity","amenity=atm","ATM","Financial"),
    ("Amenity","amenity=fuel","Gas Station","Transport"),
    ("Amenity","amenity=car_rental","Car Rental","Transport"),
    ("Amenity","amenity=bus_station","Bus Station","Transport"),
    ("Amenity","amenity=car_wash","Car Wash","Automotive"),
    ("Amenity","amenity=cinema","Cinema","Entertainment"),
    ("Amenity","amenity=theatre","Theatre","Entertainment"),
    ("Amenity","amenity=nightclub","Nightclub","Entertainment"),
    ("Amenity","amenity=marketplace","Marketplace","Retail"),
    ("Amenity","amenity=school","School","Education"),
    ("Amenity","amenity=university","University","Education"),
    ("Amenity","amenity=college","College","Education"),
    ("Amenity","amenity=kindergarten","Kindergarten","Education"),
    ("Amenity","amenity=library","Library","Education"),
    ("Amenity","amenity=gym","Gym","Health & Fitness"),
    ("Amenity","amenity=spa","Spa","Health & Fitness"),
    ("Amenity","amenity=post_office","Post Office","Government"),
    ("Amenity","amenity=police","Police","Government"),
    ("Amenity","amenity=fire_station","Fire Station","Government"),
    ("Amenity","amenity=government","Government Office","Government"),
    ("Amenity","amenity=coworking_space","Coworking Space","Business"),
    ("Amenity","amenity=community_centre","Community Centre","Social"),
    ("Shop","shop=supermarket","Supermarket","Retail"),
    ("Shop","shop=convenience","Convenience Store","Retail"),
    ("Shop","shop=bakery","Bakery","Food & Beverage"),
    ("Shop","shop=butcher","Butcher","Food & Beverage"),
    ("Shop","shop=greengrocer","Greengrocer","Food & Beverage"),
    ("Shop","shop=confectionery","Sweets","Food & Beverage"),
    ("Shop","shop=pastry","Pastry","Food & Beverage"),
    ("Shop","shop=clothes","Clothing","Retail – Fashion"),
    ("Shop","shop=shoes","Shoes","Retail – Fashion"),
    ("Shop","shop=jewelry","Jewelry","Retail – Fashion"),
    ("Shop","shop=watches","Watches","Retail – Fashion"),
    ("Shop","shop=bags","Bags","Retail – Fashion"),
    ("Shop","shop=tailor","Tailor","Retail – Fashion"),
    ("Shop","shop=fabric","Fabric","Retail – Fashion"),
    ("Shop","shop=mobile_phone","Mobile Phones","Retail – Electronics"),
    ("Shop","shop=electronics","Electronics","Retail – Electronics"),
    ("Shop","shop=computer","Computers","Retail – Electronics"),
    ("Shop","shop=furniture","Furniture","Retail – Home"),
    ("Shop","shop=hardware","Hardware","Retail – Home"),
    ("Shop","shop=car","Car Dealer","Automotive"),
    ("Shop","shop=car_repair","Car Repair","Automotive"),
    ("Shop","shop=car_parts","Car Parts","Automotive"),
    ("Shop","shop=bicycle","Bicycle Shop","Sports"),
    ("Shop","shop=books","Bookstore","Retail"),
    ("Shop","shop=stationery","Stationery","Retail"),
    ("Shop","shop=gift","Gift Shop","Retail"),
    ("Shop","shop=toys","Toys","Retail"),
    ("Shop","shop=sports","Sporting Goods","Retail"),
    ("Shop","shop=beauty","Beauty","Retail – Beauty"),
    ("Shop","shop=hairdresser","Salon","Retail – Beauty"),
    ("Shop","shop=cosmetics","Cosmetics","Retail – Beauty"),
    ("Shop","shop=perfumery","Perfume","Retail – Beauty"),
    ("Shop","shop=optician","Optician","Healthcare"),
    ("Shop","shop=florist","Florist","Retail"),
    ("Shop","shop=mall","Shopping Mall","Retail"),
    ("Shop","shop=department_store","Department Store","Retail"),
    ("Shop","shop=music","Music Store","Retail"),
    ("Shop","shop=pet","Pet Shop","Retail"),
    ("Tourism","tourism=hotel","Hotel","Accommodation"),
    ("Tourism","tourism=motel","Motel","Accommodation"),
    ("Tourism","tourism=hostel","Hostel","Accommodation"),
    ("Tourism","tourism=guest_house","Guest House","Accommodation"),
    ("Tourism","tourism=apartment","Serviced Apartment","Accommodation"),
    ("Tourism","tourism=museum","Museum","Tourism"),
    ("Tourism","tourism=attraction","Attraction","Tourism"),
    ("Tourism","tourism=gallery","Gallery","Tourism"),
    ("Tourism","tourism=viewpoint","Viewpoint","Tourism"),
    ("Tourism","tourism=camp_site","Campsite","Tourism"),
    ("Tourism","tourism=theme_park","Theme Park","Entertainment"),
    ("Tourism","tourism=zoo","Zoo","Entertainment"),
    ("Leisure","leisure=fitness_centre","Gym","Health & Fitness"),
    ("Leisure","leisure=sports_centre","Sports Center","Health & Fitness"),
    ("Leisure","leisure=stadium","Stadium","Sports"),
    ("Leisure","leisure=park","Park","Outdoor"),
    ("Leisure","leisure=garden","Garden","Outdoor"),
    ("Leisure","leisure=playground","Playground","Outdoor"),
    ("Leisure","leisure=swimming_pool","Swimming Pool","Health & Fitness"),
    ("Leisure","leisure=water_park","Water Park","Entertainment"),
    ("Leisure","leisure=golf_course","Golf Course","Sports"),
    ("Leisure","leisure=marina","Marina","Outdoor"),
    ("Leisure","leisure=dance","Dance Studio","Health & Fitness"),
    ("Leisure","leisure=bowling_alley","Bowling","Entertainment"),
    ("Leisure","leisure=nature_reserve","Nature Reserve","Outdoor"),
    ("Healthcare","healthcare=pharmacy","Pharmacy","Healthcare"),
    ("Healthcare","healthcare=hospital","Hospital","Healthcare"),
    ("Healthcare","healthcare=clinic","Clinic","Healthcare"),
    ("Healthcare","healthcare=doctor","Doctor","Healthcare"),
    ("Healthcare","healthcare=dentist","Dentist","Healthcare"),
    ("Healthcare","healthcare=laboratory","Medical Lab","Healthcare"),
    ("Healthcare","healthcare=physiotherapist","Physiotherapy","Healthcare"),
    ("Healthcare","healthcare=optometrist","Optician","Healthcare"),
    ("Office","office=estate_agent","Real Estate Agency","Business"),
    ("Office","office=insurance","Insurance","Business"),
    ("Office","office=lawyer","Law Firm","Business"),
    ("Office","office=accountant","Accounting","Business"),
    ("Office","office=government","Government Office","Government"),
    ("Office","office=travel_agent","Travel Agency","Business"),
    ("Office","office=coworking","Coworking Space","Business"),
    ("Office","office=company","Company","Business"),
    ("Office","office=it","IT Company","Technology"),
    ("Office","office=telecommunication","Telecom","Technology"),
    ("Office","office=employment_agency","Employment Agency","Business"),
]
write_sheet(ws, ["OSM Type","OSM Key=Value","Normalised Label","Group"],
            [list(r) for r in place_cat_data], widths=[12,32,24,22])
print(f"Sheet 21/24 – Place Categories ({len(place_cat_data)} types) ✓")

# ════════════════════════════════════════════════════════════════════
# SHEET 22 – GOVERNMENT INSTITUTIONS CURATED
# ════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Govt Institutions – Curated")
govt_data = [
    ("Ministry","Ministry of Finance","وزارة المالية","Fiscal policy, budget, national taxation"),
    ("Ministry","Ministry of Interior","وزارة الداخلية","Civil status, passports, residency permits"),
    ("Ministry","Ministry of Health","وزارة الصحة","Healthcare regulation, public hospitals, licensing"),
    ("Ministry","Ministry of Education","وزارة التربية والتعليم","K-12 public education curriculum & schools"),
    ("Ministry","Ministry of Higher Education & Scientific Research","وزارة التعليم العالي","Universities, research, accreditation"),
    ("Ministry","Ministry of Tourism & Antiquities","وزارة السياحة والآثار","Tourism promotion, heritage site management"),
    ("Ministry","Ministry of Trade & Industry","وزارة الصناعة والتجارة","Business registration, trade, competition policy"),
    ("Ministry","Ministry of Energy & Mineral Resources","وزارة الطاقة والثروة المعدنية","Energy, oil, gas, mining regulation"),
    ("Ministry","Ministry of Transport","وزارة النقل","Roads, aviation policy, maritime"),
    ("Ministry","Ministry of Agriculture","وزارة الزراعة","Food security, farming, water for agriculture"),
    ("Ministry","Ministry of Labor","وزارة العمل","Employment law, work permits, labour disputes"),
    ("Ministry","Ministry of Justice","وزارة العدل","Courts, judiciary, legal system"),
    ("Ministry","Ministry of Foreign Affairs","وزارة الخارجية","Diplomatic relations, visa policy, embassies"),
    ("Ministry","Ministry of Digital Economy & Entrepreneurship","وزارة الاقتصاد الرقمي","Tech policy, digital transformation, startups"),
    ("Ministry","Ministry of Water & Irrigation","وزارة المياه والري","Water resources management"),
    ("Ministry","Ministry of Environment","وزارة البيئة","Environmental regulation, nature reserves"),
    ("Ministry","Ministry of Social Development","وزارة التنمية الاجتماعية","Social welfare, NGO licensing, poverty alleviation"),
    ("Ministry","Ministry of Communication & IT","وزارة الاتصالات","Telecom regulation, internet policy"),
    ("Central / Regulatory","Central Bank of Jordan","البنك المركزي الأردني","Monetary policy, banking sector oversight, JOD"),
    ("Central / Regulatory","Jordan Securities Commission (JSC)","هيئة الأوراق المالية","Capital markets regulator"),
    ("Central / Regulatory","Amman Stock Exchange (ASE)","بورصة عمّان","National stock exchange; listed 250+ companies"),
    ("Central / Regulatory","Greater Amman Municipality","أمانة عمّان الكبرى","Amman city administration, planning, services"),
    ("Central / Regulatory","Aqaba Special Economic Zone Authority (ASEZA)","سلطة منطقة العقبة الاقتصادية الخاصة","Aqaba free zone management; tax incentives"),
    ("Central / Regulatory","Jordan Investment Commission","هيئة الاستثمار الأردنية","FDI attraction, investor services, licensing"),
    ("Central / Regulatory","Jordan Customs Department","دائرة الجمارك الأردنية","Import/export duties, border control"),
    ("Central / Regulatory","Civil Status & Passports Department","دائرة الأحوال المدنية","National IDs, passports, birth & marriage certs"),
    ("Central / Regulatory","Income & Sales Tax Department","دائرة ضريبة الدخل والمبيعات","Tax collection; GST equivalent (16%)"),
    ("Utility","National Electric Power Co. (NEPCO)","شركة الكهرباء الوطنية","National electricity transmission grid"),
    ("Utility","Water Authority of Jordan","سلطة المياه","Drinking water supply & wastewater"),
    ("Utility","Jordan Electricity Company","شركة كهرباء الأردن","Distribution in greater Amman region"),
    ("Bank","Arab Bank","البنك العربي","Largest Arab bank by assets; HQ Amman; global presence in 30+ countries"),
    ("Bank","Bank of Jordan","بنك الأردن","Retail & commercial banking"),
    ("Bank","Housing Bank for Trade & Finance","بنك الإسكان","Largest by total assets in Jordan; strong mortgage portfolio"),
    ("Bank","Jordan Ahli Bank","البنك الأهلي الأردني","Retail banking; regional network"),
    ("Bank","Cairo Amman Bank","بنك القاهرة عمّان","Retail & SME focus; 80+ branches"),
    ("Bank","Jordan Islamic Bank","البنك الإسلامي الأردني","Largest Islamic bank in Jordan; Shariah-compliant"),
    ("Bank","Safwa Islamic Bank","بنك صفوة الإسلامي","Islamic banking with growing digital capabilities"),
    ("Bank","Capital Bank","بنك كابيتال","Corporate & investment banking"),
    ("Bank","Jordan Kuwait Bank","البنك الأردني الكويتي","Retail; partially Kuwait-owned"),
    ("Bank","Arab Jordan Investment Bank (AJIB)","البنك العربي الأردني للاستثمار","Investment & corporate banking"),
    ("Digital Bank","Blink (by Cairo Amman Bank)","بلنك","Jordan's first fully digital bank; app-only"),
    ("Digital Bank","Reflect (by Bank al Etihad)","ريفلكت","Mobile-first digital banking product"),
    ("Telecom","Zain Jordan","زين الأردن","Largest operator by subscribers; 4G/5G"),
    ("Telecom","Orange Jordan","أورنج الأردن","Fixed & mobile; Orange Group subsidiary; strong fibre"),
    ("Telecom","Umniah","أمنية","Budget mobile operator; Batelco subsidiary"),
    ("Telecom","Jordan Telecom Group (JTG)","مجموعة الاتصالات الأردنية","Fixed-line incumbent; ADSL & fibre"),
    ("Transport","Royal Jordanian Airlines","الملكية الأردنية","National carrier; 50+ destinations; Oneworld member"),
    ("Transport","Queen Alia International Airport (AMM)","مطار الملكة علياء الدولي","HUB; 12M+ passengers/year; Terminals 1 & 2"),
]
write_sheet(ws, ["Sector","Institution","Arabic Name","Notes"],
            [list(r) for r in govt_data], widths=[22,50,36,58])
print(f"Sheet 22/24 – Govt Institutions Curated ({len(govt_data)} rows) ✓")

# ════════════════════════════════════════════════════════════════════
# SHEET 23 – TOURISM ATTRACTIONS CURATED
# ════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Tourism Attractions – Curated")
tourism_data = [
    ("Petra","البتراء","Ma'an","Archaeological","Yes (1985)","New Seven Wonders of the World (2007); Rose-red Nabataean city; Treasury (Al-Khazneh); Monastery (Ad Deir); Siq gorge; 800+ monuments","Mar–May, Sep–Nov","50 JOD (1 day) / 55 JOD (2 days)","Book Treasury sunrise tour; hire a guide; wear comfortable shoes; Petra by Night on Mon/Wed/Thu"),
    ("Wadi Rum","وادي رم","Aqaba","Nature / Desert","Yes (2011 – Mixed)","Red sandstone & granite desert; Bedouin culture; jeep tours; camel rides; rock climbing; stargazing; overnight camp","Mar–May, Sep–Nov","5 JOD (bus entry)","Stay overnight in Bedouin camp; book jeep tour in advance; featured in 'The Martian'"),
    ("Dead Sea (Sweimeh)","البحر الميت","Balqa","Nature / Wellness","No","World's lowest point (−430 m); 9× saltier than ocean; float naturally; mineral-rich black mud; 10+ luxury beach resorts","Mar–Jun, Sep–Nov","15–30 JOD (resort day pass)","Sunscreen essential (water is harsh); rinse mud off quickly; visit weekdays to avoid crowds"),
    ("Jerash","جرش","Jerash","Roman Archaeological","No (nominated)","'Pompeii of the East'; Hadrian's Arch; Oval Plaza; Temple of Artemis; 6,000-seat theatre; chariot racing re-enactments","All year (avoid midday Jul–Aug)","10 JOD","Jerash Festival (Jul) has live performances; 45 min from Amman"),
    ("Wadi Mujib","وادي الموجب","Madaba","Nature / Adventure","MAB Biosphere Reserve","Jordan's Grand Canyon; aquatic siq trail (wade through river); canyoning; cliff hiking; world's lowest nature reserve (−410 m)","Apr–Oct (trails closed Nov–Mar)","21 JOD (siq trail)","Book in advance via RSCN; expect to get wet; not suitable for non-swimmers"),
    ("Dana Biosphere Reserve","محمية ضانا","Tafilah","Eco-tourism / Nature","MAB Biosphere Reserve","Jordan's largest reserve (308 km²); 4 bio-geographic zones; historic Dana village; Feynan Ecolodge; 800+ plant species; Nubian ibex; sand cat","Mar–May, Sep–Nov (avoid summer heat)","Free (guide recommended)","Stay at Feynan Ecolodge for complete off-grid eco experience; amazing stargazing"),
    ("Ajloun Castle","قلعة عجلون","Ajloun","Islamic Fortress","No","12th-century fortress built by Saladin's general Izz al-Din Usama (1184–1185); panoramic views of Jordan Valley; pine forest setting","All year","3 JOD","1 hour from Amman; combine with Ajloun Forest Reserve walk"),
    ("Amman Citadel","جبل القلعة","Amman","Multi-era Archaeological","No","Inhabited since Neolithic; Temple of Hercules (2nd c. AD); Umayyad Palace (8th c.); Byzantine church; stunning Amman skyline views","All year","3.5 JOD","Free on Tuesdays; Jordan Museum is 10 min walk"),
    ("Roman Theatre, Amman","المسرح الروماني","Amman","Roman Archaeological","No","2nd-century AD; 6,000-seat capacity; built under Marcus Aurelius; still used for concerts; adjacent Odeon (500-seat)","All year","3.5 JOD","Downtown Amman; free entry on some national holidays"),
    ("Mount Nebo","جبل نيبو","Madaba","Religious / Historical","No","Where Moses viewed the Promised Land (Deuteronomy 34); Byzantine church; 4th-century mosaics; views of Dead Sea, Jordan Valley & Jerusalem on clear days","All year","3 JOD","45 min from Amman; combine with Madaba mosaic map visit"),
    ("Madaba Mosaic Map","خريطة مادبا","Madaba","Mosaic / Religious","No","World's oldest cartographic depiction of the Holy Land; 6th-century Byzantine; 2.3M tessera; St George's Church; named cities & landmarks","All year","2 JOD","Inside a functioning Greek Orthodox church; Madaba has many other mosaics"),
    ("Jordan Museum","المتحف الأردني","Amman","Museum","No","National museum; Dead Sea Scrolls; Ain Ghazal statues (7500 BC – world's oldest large-scale human figures); Roman, Islamic & Nabataean collections","All year (closed Mon)","4 JOD","Best museum in Jordan; 2–3 hours needed; near Amman Citadel"),
    ("Aqaba Marine Reserve","محمية العقبة البحرية","Aqaba","Marine / Diving","ASEZA Protected","Red Sea coral reefs; 200+ fish species; sea turtles; scuba diving; snorkelling; glass-bottom boat tours; clear water year-round (20–27°C)","Year-round (best Mar–Jun, Sep–Nov)","10 JOD (park entry)","Multiple dive centres in Aqaba city; easily reachable from town"),
    ("Baptism Site (Al-Maghtas)","الموقع الأثري المغطس","Balqa","Religious","Yes (2015)","Where Jesus was baptised by John the Baptist; Jordan River; Byzantine churches; ongoing excavations; pilgrimage site for Christians worldwide","All year","12 JOD","45 min from Amman; coordinate with nearby Dead Sea visit"),
    ("Azraq Wetland Reserve","محمية الأزرق الرطبة","Zarqa","Nature / Birdwatching","Ramsar Wetland","Oasis in eastern desert; 300+ bird species; migratory birds; Lawrence of Arabia HQ in WWI; rare White Pelican; Black Iris (Jordan's national flower)","Oct–Apr (peak migration)","3 JOD","RSCN managed; 90 min from Amman; combine with Azraq Castle visit"),
]
write_sheet(ws, ["Site","Arabic Name","Governorate","Type","UNESCO","Highlights","Best Time to Visit","Entry Fee","Tips"],
            [list(r) for r in tourism_data],
            widths=[24,22,12,22,12,68,26,22,52])
print(f"Sheet 23/24 – Tourism Attractions ({len(tourism_data)} rows) ✓")

# ════════════════════════════════════════════════════════════════════
# SHEET 24 – MAJOR CORPORATIONS CURATED
# ════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Major Corporations – Curated")
corp_data = [
    ("Tech / Internet","Maktoob","1999","Acquired by Yahoo ($165M, 2009)","Pioneer Arab internet portal; spawned Jordan's startup ecosystem; email, news, entertainment"),
    ("Tech / E-Commerce","Souq.com","2005","Acquired by Amazon (~$650M, 2017)","Arab world's largest e-commerce marketplace; co-founded by Jordanian Ronaldo Mouchawar; became Amazon.ae"),
    ("Tech / Books","Jamalon","2010","Active – Private","Arab world's largest online bookstore; HQ Amman; 10M+ Arabic & English titles; 22+ Arab countries"),
    ("Tech / Gaming","Tamatem Games","2013","Active – Private","Mobile gaming for Arabic market; 25M+ downloads; HQ Amman; localises global games"),
    ("Tech / Content","Mawdoo3","2011","Active – Private","Largest Arabic content platform; Wikipedia-like; 100M+ monthly visitors; 1,500+ content creators"),
    ("Tech / Animation","Kharabeesh","2007","Active – Private","Leading Arabic animation & digital content studio; YouTube presence; political satire"),
    ("Tech / Logistics","Aramex","1982","Public – ADX (ARMX)","Global logistics & courier; founded Amman by Fadi Ghandour; listed Abu Dhabi; 600+ cities; Jordan's first startup-to-global-company story"),
    ("Tech / Payments","eFAWATEERcom","2014","Active (CBJ)","Jordan's national bill payment platform; Central Bank backed; 700+ billers; 40+ bank integrations"),
    ("Tech / Accelerator","Oasis500","2010","Active","Jordan's leading startup accelerator; 200+ startups funded; backed by King Abdullah II"),
    ("Tech / Accelerator","Endeavor Jordan","2004","Active","Global high-impact entrepreneur network; mentorship & funding for scale-up companies"),
    ("Pharma","Hikma Pharmaceuticals","1978","Public – LSE (HIK)","LSE-listed global pharma; founded Amman by Samih Darwazah; generics & injectables; 30+ countries; revenues $3B+"),
    ("Pharma","Jordan Pharmaceutical Manufacturing (JPM)","1978","Active – Private","Leading Jordanian generic drug manufacturer; GMP certified; exports to 60+ countries"),
    ("Healthcare","King Hussein Cancer Center (KHCC)","2003","Public Institution","Jordan's premier cancer care; JCI accredited (2006, first in Arab world); serves patients from 70+ countries"),
    ("Healthcare","Royal Medical Services (RMS)","1950","Government","Military & government healthcare network; largest by bed count in Jordan; full range of specialties"),
    ("Mining","Jordan Phosphate Mines Co. (JPMC)","1953","Public – ASE (JOPH)","3rd largest phosphate producer globally; 20M+ tons/year; major national export earner"),
    ("Mining","Arab Potash Company (APC)","1956","Public – ASE (APOT)","Extracts potash from Dead Sea; 2.5M ton/year; major global potash supplier; Dead Sea concession"),
    ("Energy","Jordan Petroleum Refinery Co. (JPRC)","1956","Public","Sole petroleum refinery in Jordan; Zarqa; processes 100,000 bbl/day; supplies most of Jordan's fuel"),
    ("Supermarket","Carrefour Jordan (Majid Al Futtaim)","1999","Active – Franchise","10+ hypermarkets across Jordan; dominant grocery chain; anchor tenant at City Mall, Mecca Mall"),
    ("Supermarket","Safeway Jordan","1999","Active","Premium supermarket; 5+ west Amman branches; strong in expat communities"),
    ("Supermarket","Cozmo (Lemongrass Group)","2010","Active","Upscale grocery & gourmet; 10+ branches; trendy west Amman; strong fresh & international section"),
    ("Retail Mall","City Mall","2005","Active","Largest mall in Jordan; Mecca Street; 350+ stores; IKEA anchor; Geant hypermarket"),
    ("Retail Mall","Mecca Mall","1999","Active","Pioneer enclosed mall; Sweifieh; 220+ stores; first major mall in Jordan"),
    ("Retail Mall","Taj Mall","2008","Active","Luxury mall; 7th Circle; premium brands; luxury cinemas; 230+ stores"),
    ("Retail Mall","Abdali Mall","2016","Active","Modern downtown Abdali development; sky bridges connecting buildings; 350+ stores"),
    ("Retail Mall","Avenue Mall","2021","Active","Newest large-format mall; South Amman; family-oriented; open-air sections"),
    ("Food Delivery","Talabat Jordan","2004","Active (Delivery Hero)","Dominant food delivery app; 3,000+ restaurant partners; grocery delivery; operates across MENA"),
    ("Ride-Hailing","Careem / Uber Jordan","2015","Active","Careem acquired by Uber 2019 for $3.1B; both brands operate in Jordan; dominant ride-hailing"),
    ("Media","Roya TV","2010","Active","Leading private TV station; most-watched in Jordan; news, entertainment, drama"),
    ("Media","Jordan Times","1975","Active","Jordan's oldest English-language daily newspaper; diplomatic & business community"),
    ("Media","Al-Ghad","2004","Active","Leading Arabic-language daily; among top online news sites in Jordan"),
    ("Airlines","Royal Jordanian (RJ)","1963","Public – ASE","National carrier; Oneworld member; 50+ destinations; HUB at Amman (AMM)"),
    ("Logistics","JETT (Jordan Express Tourist Transport)","1972","Government","National bus company; intercity routes; also operates tourist transport"),
    ("Startup","Rubicon","2001","Active","Specialised Jordanian tech consultancy; AI & digital transformation; works with Gulf governments"),
]
write_sheet(ws, ["Sector","Company","Founded","Status","Description"],
            [list(r) for r in corp_data], widths=[22,34,10,28,70])
print(f"Sheet 24/24 – Major Corporations ({len(corp_data)} rows) ✓")

# ── Save ─────────────────────────────────────────────────────────────────────
print(f"\nSaving to {OUT}…", end=" ", flush=True)
wb.save(OUT)
print("Done.")
print(f"\n✓ {OUT}")
print(f"  {len(wb.sheetnames)} sheets: {', '.join(wb.sheetnames)}")
